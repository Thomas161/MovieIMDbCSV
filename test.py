from openpyxl import Workbook
from imdb import Cinemagoer
from openpyxl.styles import Alignment, Font, PatternFill
# from openpyxl import Workbook, load_workbook
im = Cinemagoer()
wb = Workbook()

ws = wb.create_sheet("Sheet_A")
ws.title = "Films"

center_align = Alignment(horizontal='center', vertical='center')

ws['B1'] = "Title"
ws['B1'].font = Font(name='Verdana', size=18, bold=True, color='00FFFFFF')
ws['B1'].fill = PatternFill("solid", start_color='00339966')
ws['C1'] = "Director"
ws['C1'].font = Font(name='Verdana', size=18, bold=True, color='00FFCC00')
ws['C1'].fill = PatternFill("solid", start_color='00FF8080')
ws['D1'] = "Year"
ws['D1'].font = Font(name='Verdana', size=18, bold=True, color='00FF6600')
ws['D1'].fill = PatternFill("solid", start_color='00003366')
ws['E1'] = "Genre"
ws['E1'].font = Font(name='Verdana', size=18, bold=True, color='000066CC')
ws['E1'].fill = PatternFill("solid", start_color='00FFFF00')

for c in ws['A2:A100']:
    c[0].alignment = center_align
for c in ws['B2:D100']:
    c[0].font = Font(size=16, italic=True)

for i in range(1, 100):
    ws.cell(row=i+1, column=1, value="\*")

films = ["0078748", "0093773", "1375666",
         "0090605", "0167261", "0369339",
         "0372784", "0468569", "0080455",
         "0167260", "0120737", "0208092",
         "0253556", "0238380", "0209144",
         "0278504", "0076740", "0090180",
         "0147800", "0112864", "0114369",
         "0113189", "0381061", "0075005",
         "1853728", "0110912", "0361748",
         "1392214", "1856101", "3397884",
         "0119081", "0096256", "0107076",
         "0364725", "0172495", "0265086",
         "0190590", "0116282", "0108358",
         "0058461", "0059578", "1205489",
         "0103644", "0396269", "0374900",
         "0083658", "0103064", "0191397",
         "0118887", "0107290", "0418279",
         "0066999", "2265171", "0120611",
         "1745960", "0187078", "0246578",
         "0110413", "0116483", "0377092",
         "0166924", "0086190", "1049413",
         "0129167", "0120363", "0120755",
         "4912910", "0076729", "0067116",
         "0348333", "0115759", "0109040",
         "0113277", "0947810", "0096874",
         "0320661", "0097733", "0477348",
         "0082869", "0108399", "0104348",
         "0120201", "0075784", "0120863",
         "0118880", "0211915", "0100403",
         "0084434", "0117998", "0079817",
         "0469494", "0146838", "0044079",
         "0102138", "0104684", "0100802",
         "0120586", "0085636", "10045260"]

column = 2
columnThree = 3
columnFour = 4
columnFive = 5


def render_films():
    for i, v in enumerate(films):
        movie = im.get_movie(v)
        title = movie['original title']
        year = movie['year']
        genre = movie['genres']
        for e in genre:
            for d in movie['directors']:
                ws.cell(row=i+2, column=column, value=title)
                ws.cell(row=i+2, column=columnThree, value=str(d['name']))
                ws.cell(row=i+2, column=columnFour, value=year)
                ws.cell(row=i+2, column=columnFive, value=e)


render_films()
wb.save('films.xlsx')
